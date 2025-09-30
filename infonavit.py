import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import re
import openpyxl
from openpyxl.styles import PatternFill

st.title("Limpieza y comparación de archivo Excel")

# ------------------- Carga y limpieza de SICOSS -------------------
archivo_base = st.file_uploader("Carga tu archivo Excel de SICOSS", type=["xlsx"], key="base")
preview_data = pd.DataFrame()  # Inicializar variable para evitar errores

if archivo_base:
    df = pd.read_excel(archivo_base, engine="openpyxl")
    header_rows = df.iloc[:5]
    data_rows = df.iloc[5:]

    exclusion_keywords = ["Total Trabajadores", "CENTRO DE TRABAJO"]
    mask_keywords = ~data_rows.apply(
        lambda row: any(keyword in str(cell) for cell in row for keyword in exclusion_keywords),
        axis=1
    )
    filtered_data = data_rows[mask_keywords].dropna(how='all')

    if filtered_data.shape[1] > 1:
        filtered_data.loc[:, filtered_data.columns[1]] = filtered_data.iloc[:, 1].astype(str).str.replace("-", "", regex=False)

    final_df = pd.concat([header_rows, filtered_data], ignore_index=True)

    preview_data = filtered_data.copy()
    preview_data.columns = df.iloc[4].fillna("").astype(str).str.strip()
    preview_data = preview_data.dropna(how='all')

    if preview_data.shape[1] > 1:
        preview_data.loc[:, preview_data.columns[1]] = preview_data.iloc[:, 1].astype(str).str.replace("-", "", regex=False)
    if "TIPO DE DESCUENTO" in preview_data.columns:
        preview_data["TIPO DE DESCUENTO"] = preview_data["TIPO DE DESCUENTO"].astype(str).apply(
            lambda x: re.search(r'\d', x).group(0) if re.search(r'\d', x) else "0"
        )

    st.write("Vista previa de SICOSS:")
    st.dataframe(preview_data)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        final_df.to_excel(writer, index=False)
    output.seek(0)
    st.download_button(
        label="Descargar archivo limpio",
        data=output,
        file_name="archivo_limpio.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ------------------- Comparación con INFONAVIT -------------------
archivo_comp = st.file_uploader("Carga el archivo de INFONAVIT", type=["xlsx"], key="comparacion")

if archivo_comp and not preview_data.empty:
    df_temp = pd.read_excel(archivo_comp, engine="openpyxl", header=None)
    columnas_esperadas = ["NSS", "Número de Crédito", "Tipo de descuento", "Valor de descuento"]
    encabezado_index = None
    for i in range(5):
        if all(col in df_temp.iloc[i].astype(str).str.strip().tolist() for col in columnas_esperadas):
            encabezado_index = i
            break
    if encabezado_index is None:
        st.error("No se pudo detectar automáticamente el encabezado del archivo de comparación.")
        st.stop()

    df_comp = pd.read_excel(archivo_comp, engine="openpyxl", header=encabezado_index, dtype={"NSS": str})
    df_comp.columns = df_comp.columns.str.strip()
    df_comp = df_comp.dropna(how='all')
    df_comp["NSS"] = df_comp["NSS"].str.zfill(11)

    if "Tipo de descuento" in df_comp.columns:
        df_comp["Tipo de descuento"] = df_comp["Tipo de descuento"].replace({
            "4": "2",
            "6": "3",
            4: "2",
            6: "3"
        })

    st.success(f"Encabezado detectado en fila {encabezado_index + 1}")

    comparacion_base = preview_data.rename(columns={
        "NUM. SEGURIDAD SOCIAL": "NSS",
        "NUM. CREDITO INFONAVIT": "Número de Crédito",
        "TIPO DE DESCUENTO": "Tipo de descuento",
        "VALOR DE DESCUENTO": "Valor de descuento"
    })
    columnas_comparacion = ["NSS", "Número de Crédito", "Tipo de descuento", "Valor de descuento"]

    base_final = comparacion_base[columnas_comparacion].copy()
    comp_final = df_comp[columnas_comparacion].copy()
    base_final["NSS"] = base_final["NSS"].astype(str).str.strip().str.zfill(11)
    comp_final["NSS"] = comp_final["NSS"].astype(str).str.strip().str.zfill(11)

    # ------------------ FILTRAR SOLO NSS EXISTENTES EN INFONAVIT ------------------
    base_final = base_final[base_final["NSS"].isin(comp_final["NSS"])].copy()
    # ------------------------------------------------------------------------------

    # ------------------ NORMALIZAR COLUMNAS ANTES DEL MERGE ------------------
    def normalizar_columna(serie, tipo="texto"):
        if tipo == "credito":
            return serie.fillna("0").astype(str).str.strip().replace(["nan","NaN","None",""],"0").apply(
                lambda x: str(int(x)) if x.isdigit() else x.lstrip("0") if x.replace(".", "", 1).isdigit() else x
            )
        elif tipo == "descuento":
            return serie.fillna("0").astype(str).str.strip().replace(["nan","NaN","None",""],"0").apply(
                lambda x: str(int(float(x))) if x.replace(".", "", 1).isdigit() else "0"
            ).replace({"4": "2", "6": "3"})
        elif tipo == "valor":
            return serie.fillna("0").astype(str).str.strip().replace(["nan","NaN","None",""],"0").apply(
                lambda x: f"{float(x):.2f}" if x.replace(".", "", 1).isdigit() else "0"
            )
        else:
            return serie.fillna("0").astype(str).str.strip().replace(["nan","NaN","None",""],"0")

    base_final["Número de Crédito"] = normalizar_columna(base_final["Número de Crédito"], "credito")
    comp_final["Número de Crédito"] = normalizar_columna(comp_final["Número de Crédito"], "credito")
    base_final["Tipo de descuento"] = normalizar_columna(base_final["Tipo de descuento"], "descuento")
    comp_final["Tipo de descuento"] = normalizar_columna(comp_final["Tipo de descuento"], "descuento")
    base_final["Valor de descuento"] = normalizar_columna(base_final["Valor de descuento"], "valor")
    comp_final["Valor de descuento"] = normalizar_columna(comp_final["Valor de descuento"], "valor")
    # ------------------------------------------------------------------------------

    # ------------------ RECORTAR DUPLICADOS EN SICOSS SI SOBRAN ------------------
    counts_infonavit = comp_final.groupby("NSS").size()
    def limitar_duplicados(group):
        n_max = counts_infonavit.get(group.name, 0)
        return group.head(n_max)
    base_final = base_final.groupby("NSS", group_keys=False).apply(limitar_duplicados)
    # ------------------------------------------------------------------------------

    # ------------------ CREAR INDICE PARA ALINEAR DUPLICADOS ------------------
    base_final["dup_idx"] = base_final.groupby("NSS").cumcount()
    comp_final["dup_idx"] = comp_final.groupby("NSS").cumcount()

    merged = pd.merge(
        base_final,
        comp_final,
        on=["NSS", "dup_idx"],
        how="outer",
        suffixes=("_SICOSS", "_INFONAVIT"),
        indicator=True
    )
    merged.drop(columns=["dup_idx"], inplace=True)
    # ------------------------------------------------------------------------------

    cols = ["Número de Crédito", "Tipo de descuento", "Valor de descuento"]

    detalles_list = []
    estados = []
    for idx in range(len(merged)):
        detalles = []
        for c in cols:
            if merged.at[idx, f"{c}_SICOSS"] != merged.at[idx, f"{c}_INFONAVIT"]:
                detalles.append(f"{c}: SICOSS='{merged.at[idx, f'{c}_SICOSS']}' vs INFONAVIT='{merged.at[idx, f'{c}_INFONAVIT']}'")
        detalle_str = "; ".join(detalles)
        detalles_list.append(detalle_str)
        estados.append("DIFIERE" if detalle_str else "NO DIFIERE")

    merged["Detalle"] = detalles_list
    merged["Estado"] = estados

    def resaltar_diferencias(row):
        estilos = [""] * len(merged.columns)
        for col in cols:
            if row[f"{col}_SICOSS"] != row[f"{col}_INFONAVIT"]:
                estilos[merged.columns.get_loc(f"{col}_SICOSS")] = "background-color: #FF9999"
                estilos[merged.columns.get_loc(f"{col}_INFONAVIT")] = "background-color: #FF9999"
        return estilos

    st.write("Comparación completa:")
    st.dataframe(merged.style.apply(resaltar_diferencias, axis=1))

    output_dif = BytesIO()
    with pd.ExcelWriter(output_dif, engine='openpyxl') as writer:
        merged.to_excel(writer, index=False, sheet_name="Comparacion")
        ws = writer.sheets["Comparacion"]
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        for row_idx in range(len(merged)):
            for col in cols:
                if merged.at[row_idx, f"{col}_SICOSS"] != merged.at[row_idx, f"{col}_INFONAVIT"]:
                    ws.cell(row=row_idx+2, column=merged.columns.get_loc(f"{col}_SICOSS")+1).fill = red_fill
                    ws.cell(row=row_idx+2, column=merged.columns.get_loc(f"{col}_INFONAVIT")+1).fill = red_fill

    output_dif.seek(0)
    st.download_button(
        label="Descargar comparación resaltada en Excel",
        data=output_dif,
        file_name="comparacion_resaltada.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
